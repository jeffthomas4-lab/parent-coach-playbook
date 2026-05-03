#!/usr/bin/env python3
r"""
import-camps.py

Reads the filled camps-template.xlsx (or any CSV with the same headers),
validates each row, geocodes the address via Nominatim, generates a UUID
and unique slug, and writes a SQL file of INSERT statements against the
parent-coach-playbook D1 database.

All inserts default to status='approved', verified=1, submitted by Jeff,
auto-reviewed at run time. Run the resulting SQL with wrangler:

    npx wrangler d1 execute parent-coach-playbook ^
        --file=imports/out/camps-import-YYYY-MM-DD.sql --remote

Usage (PowerShell):

    cd "C:\Users\jeffthomas\Desktop\Claude Cowork\Outputs\parent-coach-playbook"
    python imports/import-camps.py imports/camps-template.xlsx

Optional flags:
    --no-geocode        Skip Nominatim. Latitude/longitude written as NULL.
    --status pending    Override status (default: approved)
    --submitter EMAIL   Override submitter email (default: jeffthomas@pugetsound.edu)
    --out PATH          Override output SQL path
    --skipped PATH      Path to write a CSV of rows that failed validation
    --anchor TEXT       Anchor area for this batch (e.g., 'Tacoma, WA (25mi)').
                        When set, a row is appended to CAMP-SEARCH-LOG.md.
    --no-log            Skip the CAMP-SEARCH-LOG.md append even if --anchor is set.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
import time
import urllib.parse
import urllib.request
import uuid
from pathlib import Path
from typing import Any

DEFAULT_SUBMITTER = "jeffthomas@pugetsound.edu"
DEFAULT_REVIEWER = "jeff (bulk import)"
NOMINATIM_USER_AGENT = (
    "parentcoachplaybook.com camps directory (parentcoachplaybook@gmail.com)"
)
NOMINATIM_DELAY_S = 1.1

REQUIRED_FIELDS = [
    "name", "sport", "age_min", "age_max", "start_date", "end_date",
    "address", "city", "state", "zip", "description",
]

ALL_COLUMNS = REQUIRED_FIELDS + [
    "price_text", "day_or_overnight", "skill_level", "spots_status",
    "contact_email", "contact_phone", "website_url",
    "lunch_included", "aftercare_available",
    "program_type", "registration_deadline", "schedule_text",
]

DAY_OVERNIGHT = {"day", "overnight"}
SKILL_LEVELS = {"beginner", "intermediate", "advanced", "all"}
SPOTS_STATUS = {"open", "waitlist", "full"}
PROGRAM_TYPES = {"camp", "league"}

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
ZIP_RE = re.compile(r"^\d{5}(-\d{4})?$")


def sql_str(v: Any) -> str:
    if v is None or v == "":
        return "NULL"
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return "NULL"
        return repr(v)
    s = str(v).replace("'", "''")
    return f"'{s}'"


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"^-+|-+$", "", s)
    return s[:80] or "camp"


def parse_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in {"true", "1", "yes", "y", "on"}


def parse_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def normalize_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    if DATE_RE.match(s):
        return s
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_zip(v: Any) -> str | None:
    if v is None or v == "":
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if s.isdigit() and len(s) < 5:
        s = s.zfill(5)
    if ZIP_RE.match(s):
        return s
    return None


def normalize_url(v: Any) -> str | None:
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    if not re.match(r"^https?://", s, re.IGNORECASE):
        s = f"https://{s}"
    return s


def read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            sys.exit(f"openpyxl is required. pip install openpyxl. ({e})")
        wb = load_workbook(path, data_only=True)
        ws = wb["Camps"] if "Camps" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        header = [str(h).strip() if h is not None else "" for h in header]
        out = []
        for row in rows_iter:
            record = {header[i]: row[i] for i in range(len(header))}
            if not any(v not in (None, "") for v in record.values()):
                continue
            out.append(record)
        return out
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [dict(r) for r in csv.DictReader(f)]
    sys.exit(f"unsupported input file type: {suffix}")


class RowError(Exception):
    pass


def validate_row(row: dict[str, Any], idx: int) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for f in REQUIRED_FIELDS:
        v = row.get(f)
        if v is None or str(v).strip() == "":
            raise RowError(f"row {idx}: missing required field '{f}'")

    name = str(row["name"]).strip()
    if len(name) > 200:
        raise RowError(f"row {idx}: name longer than 200 chars")
    cleaned["name"] = name
    cleaned["sport"] = str(row["sport"]).strip()

    age_min = parse_int(row["age_min"])
    age_max = parse_int(row["age_max"])
    if age_min is None or age_max is None:
        raise RowError(f"row {idx}: age_min and age_max must be integers")
    if age_min < 3 or age_max > 22:
        raise RowError(f"row {idx}: ages out of range (3-22)")
    if age_min > age_max:
        raise RowError(f"row {idx}: age_min > age_max")
    cleaned["age_min"] = age_min
    cleaned["age_max"] = age_max

    start = normalize_date(row["start_date"])
    end = normalize_date(row["end_date"])
    if not start or not end:
        raise RowError(f"row {idx}: start_date and end_date must be YYYY-MM-DD")
    if start > end:
        raise RowError(f"row {idx}: start_date after end_date")
    cleaned["start_date"] = start
    cleaned["end_date"] = end

    cleaned["address"] = str(row["address"]).strip()
    cleaned["city"] = str(row["city"]).strip()
    cleaned["state"] = str(row["state"]).strip().upper()[:2]

    z = normalize_zip(row["zip"])
    if not z:
        raise RowError(f"row {idx}: zip is not 5 digits")
    cleaned["zip"] = z

    desc = str(row["description"]).strip()
    if len(desc) < 30:
        raise RowError(f"row {idx}: description shorter than 30 chars")
    if len(desc) > 4000:
        raise RowError(f"row {idx}: description longer than 4000 chars")
    cleaned["description"] = desc

    price = row.get("price_text")
    cleaned["price_text"] = str(price).strip() if price not in (None, "") else None

    day = (str(row.get("day_or_overnight") or "day")).strip().lower()
    if day not in DAY_OVERNIGHT:
        raise RowError(f"row {idx}: day_or_overnight must be 'day' or 'overnight'")
    cleaned["day_or_overnight"] = day

    skill = (str(row.get("skill_level") or "all")).strip().lower()
    if skill not in SKILL_LEVELS:
        raise RowError(f"row {idx}: skill_level invalid")
    cleaned["skill_level"] = skill

    spots = (str(row.get("spots_status") or "open")).strip().lower()
    if spots not in SPOTS_STATUS:
        raise RowError(f"row {idx}: spots_status invalid")
    cleaned["spots_status"] = spots

    ce = row.get("contact_email")
    ce = str(ce).strip() if ce not in (None, "") else None
    if ce and not EMAIL_RE.match(ce):
        raise RowError(f"row {idx}: contact_email is not a valid email")
    cleaned["contact_email"] = ce

    cp = row.get("contact_phone")
    cleaned["contact_phone"] = str(cp).strip() if cp not in (None, "") else None
    cleaned["website_url"] = normalize_url(row.get("website_url"))
    cleaned["lunch_included"] = parse_bool(row.get("lunch_included"))
    cleaned["aftercare_available"] = parse_bool(row.get("aftercare_available"))

    program_type = (str(row.get("program_type") or "camp")).strip().lower()
    if program_type not in PROGRAM_TYPES:
        raise RowError(f"row {idx}: program_type must be 'camp' or 'league'")
    cleaned["program_type"] = program_type

    rd = row.get("registration_deadline")
    if rd in (None, ""):
        cleaned["registration_deadline"] = None
    else:
        normalized = normalize_date(rd)
        if not normalized:
            raise RowError(f"row {idx}: registration_deadline must be YYYY-MM-DD or blank")
        cleaned["registration_deadline"] = normalized

    sched = row.get("schedule_text")
    cleaned["schedule_text"] = str(sched).strip() if sched not in (None, "") else None

    return cleaned


def geocode(address: str, city: str, state: str, zipc: str):
    q = f"{address}, {city}, {state} {zipc}"
    params = urllib.parse.urlencode({"q": q, "format": "json", "limit": "1"})
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    req = urllib.request.Request(url, headers={
        "User-Agent": NOMINATIM_USER_AGENT,
        "Accept-Language": "en",
    })
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"  geocode failed for {q!r}: {e}", file=sys.stderr)
        return None
    if not data:
        return None
    try:
        return float(data[0]["lat"]), float(data[0]["lon"])
    except (KeyError, ValueError, TypeError):
        return None


INSERT_COLUMNS = [
    "id", "slug", "name", "sport", "age_min", "age_max", "start_date", "end_date",
    "address", "city", "state", "zip", "latitude", "longitude",
    "description", "price_text", "day_or_overnight", "skill_level", "spots_status",
    "contact_email", "contact_phone", "website_url",
    "lunch_included", "aftercare_available",
    "status", "submitted_by_email", "submitted_at",
    "reviewed_by", "reviewed_at", "verified",
    "program_type", "registration_deadline", "schedule_text",
]


def render_insert(rec: dict[str, Any]) -> str:
    # Use INSERT OR IGNORE so re-runs are safe: rows whose slug already exists
    # in D1 are silently skipped instead of aborting the whole batch.
    parts = [sql_str(rec[c]) for c in INSERT_COLUMNS]
    cols = ", ".join(INSERT_COLUMNS)
    vals = ", ".join(parts)
    return f"INSERT OR IGNORE INTO camps ({cols}) VALUES ({vals});"


def append_search_log(
    log_path: Path,
    today: str,
    anchor: str,
    source_name: str,
    imported: int,
    skipped: int,
    notes: str = "",
) -> bool:
    """Append a row to CAMP-SEARCH-LOG.md Batch History. Update Last updated.

    Returns True if the file was modified.
    """
    if not log_path.exists():
        print(f"  CAMP-SEARCH-LOG.md not found at {log_path}; skipping log update")
        return False
    content = log_path.read_text(encoding="utf-8")

    # Update the 'Last updated' line near the top.
    content = re.sub(
        r"(\*\*Last updated:\*\*\s*)\d{4}-\d{2}-\d{2}",
        rf"\g<1>{today}",
        content,
        count=1,
    )

    marker = "## Batch History"
    idx = content.find(marker)
    if idx < 0:
        print(f"  no '## Batch History' section in {log_path.name}; skipping log update")
        return False

    section_end = content.find("\n## ", idx + len(marker))
    if section_end < 0:
        section_end = len(content)

    notes_safe = notes.replace("|", "\\|") if notes else ""
    new_row = f"| {today} | {anchor} | {source_name} | {imported} | {skipped} | {notes_safe} |"

    section = content[idx:section_end].rstrip()
    new_section = section + "\n" + new_row + "\n"
    new_content = content[:idx] + new_section + content[section_end:]
    log_path.write_text(new_content, encoding="utf-8")
    print(f"  appended batch row to {log_path.name}")
    return True


def unique_slug(name: str, used: set) -> str:
    base = slugify(name)
    if base not in used:
        used.add(base)
        return base
    for _ in range(50):
        cand = f"{base}-{uuid.uuid4().hex[:4]}"[:80]
        if cand not in used:
            used.add(cand)
            return cand
    cand = f"{base}-{uuid.uuid4().hex[:8]}"[:80]
    used.add(cand)
    return cand


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("input")
    p.add_argument("--no-geocode", action="store_true")
    p.add_argument("--status", default="approved", choices=["approved", "pending"])
    p.add_argument("--submitter", default=DEFAULT_SUBMITTER)
    p.add_argument("--reviewer", default=DEFAULT_REVIEWER)
    p.add_argument("--out", default=None)
    p.add_argument("--skipped", default=None)
    p.add_argument("--anchor", default=None,
                   help="Anchor area for this batch (e.g., 'Tacoma, WA (25mi)'). "
                        "When set, appends a row to CAMP-SEARCH-LOG.md.")
    p.add_argument("--no-log", action="store_true",
                   help="Skip the CAMP-SEARCH-LOG.md append even if --anchor is set.")
    args = p.parse_args()

    in_path = Path(args.input).resolve()
    if not in_path.exists():
        sys.exit(f"input not found: {in_path}")

    rows = read_rows(in_path)
    if not rows:
        sys.exit("input has no data rows")

    valid = []
    skipped = []

    print(f"read {len(rows)} rows from {in_path.name}")

    for i, row in enumerate(rows, start=2):
        nm = str(row.get("name") or "").strip().upper()
        if nm.startswith("EXAMPLE"):
            print(f"  SKIP row {i}: example row (delete it from the spreadsheet to silence this)")
            skipped.append((i, row, "example row - not imported"))
            continue
        try:
            cleaned = validate_row(row, i)
        except RowError as e:
            print(f"  SKIP {e}")
            skipped.append((i, row, str(e)))
            continue
        valid.append({"_row": i, **cleaned})

    print(f"{len(valid)} valid, {len(skipped)} skipped")
    if not valid:
        sys.exit("nothing to import")

    if args.no_geocode:
        print("--no-geocode set: latitude/longitude will be NULL")
        for rec in valid:
            rec["latitude"] = None
            rec["longitude"] = None
    else:
        seen = {}
        for rec in valid:
            key = f"{rec['address']}|{rec['city']}|{rec['state']}|{rec['zip']}".lower()
            if key in seen:
                geo = seen[key]
            else:
                print(f"  geocoding: {rec['address']}, {rec['city']}, {rec['state']} {rec['zip']}")
                geo = geocode(rec["address"], rec["city"], rec["state"], rec["zip"])
                seen[key] = geo
                time.sleep(NOMINATIM_DELAY_S)
            if geo:
                rec["latitude"], rec["longitude"] = geo
            else:
                rec["latitude"], rec["longitude"] = None, None

    now_iso = dt.datetime.now(dt.timezone.utc).isoformat()
    used_slugs = set()
    for rec in valid:
        rec["id"] = str(uuid.uuid4())
        rec["slug"] = unique_slug(rec["name"], used_slugs)
        rec["status"] = args.status
        rec["submitted_by_email"] = args.submitter.lower()
        rec["submitted_at"] = now_iso
        rec["reviewed_by"] = args.reviewer if args.status == "approved" else None
        rec["reviewed_at"] = now_iso if args.status == "approved" else None
        rec["verified"] = 1 if args.status == "approved" else 0

    if args.out:
        out_path = Path(args.out).resolve()
    else:
        out_dir = in_path.parent / "out"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"camps-import-{dt.date.today().isoformat()}.sql"

    approved_n = len(valid) if args.status == "approved" else 0
    submitter_email = args.submitter.lower()
    submitter_insert = (
        "INSERT INTO submitters (email, trust_level, submission_count, approved_count, "
        "first_submitted_at, last_submitted_at, notes) VALUES ("
        f"{sql_str(submitter_email)}, 'trusted', {len(valid)}, {approved_n}, "
        f"{sql_str(now_iso)}, {sql_str(now_iso)}, 'bulk import') "
        "ON CONFLICT(email) DO UPDATE SET "
        f"submission_count = submission_count + {len(valid)}, "
        f"approved_count = approved_count + {approved_n}, "
        "last_submitted_at = excluded.last_submitted_at;"
    )

    # Cloudflare D1 over wrangler does not accept BEGIN TRANSACTION / COMMIT.
    # Each statement is executed independently. We just emit them in order.
    lines = [
        f"-- Generated {now_iso}",
        f"-- Source: {in_path.name}",
        f"-- {len(valid)} camps, status={args.status}, submitter={args.submitter}",
        "",
        "-- Make sure the submitter exists.",
        submitter_insert,
        "",
    ]
    for rec in valid:
        lines.append(render_insert(rec))
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {out_path} ({len(valid)} INSERTs)")

    if skipped:
        skipped_path = Path(args.skipped) if args.skipped else out_path.with_name(out_path.stem + "-skipped.csv")
        with skipped_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["row_number", "reason"] + ALL_COLUMNS)
            for idx, raw, reason in skipped:
                writer.writerow([idx, reason] + [raw.get(c, "") for c in ALL_COLUMNS])
        print(f"wrote {skipped_path} ({len(skipped)} skipped rows)")

    if args.no_log:
        print("--no-log set: CAMP-SEARCH-LOG.md not updated")
    elif not args.anchor:
        print("note: --anchor not provided; CAMP-SEARCH-LOG.md not updated.")
        print("      pass --anchor 'Tacoma, WA (25mi)' (or similar) to log this batch.")
    else:
        log_path = in_path.parent / "CAMP-SEARCH-LOG.md"
        append_search_log(
            log_path=log_path,
            today=dt.date.today().isoformat(),
            anchor=args.anchor,
            source_name=in_path.name,
            imported=len(valid),
            skipped=len(skipped),
            notes="",
        )

    print("\nNext step:")
    print(f'  npx wrangler d1 execute parent-coach-playbook --file="{out_path}" --remote')
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
